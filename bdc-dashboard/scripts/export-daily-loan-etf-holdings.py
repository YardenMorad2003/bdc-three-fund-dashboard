from __future__ import annotations

import hashlib
import html
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.request import Request, urlopen

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_PATH = PROJECT_ROOT / "lib" / "daily-loan-etf-holdings.json"
ARCHIVE_ROOT = PROJECT_ROOT / "public" / "data" / "loan-etf-holdings"

SRLN_PAGE = "https://www.ssga.com/us/en/individual/etfs/state-street-blackstone-senior-loan-etf-srln"
SRLN_DOWNLOAD = "https://www.ssga.com/library-content/products/fund-data/etfs/us/holdings-daily-us-en-srln.xlsx"
BKLN_PAGE = "https://www.invesco.com/us/en/financial-products/etfs/invesco-senior-loan-etf.html"
BKLN_API = (
    "https://dng-api.invesco.com/cache/v1/accounts/en_US/shareclasses/46138G508/holdings/fund"
    "?idType=cusip&productType=ETF"
)

REQUEST_HEADERS = {
    "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/126 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
}


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def fetch_bytes(url: str, *, referer: str | None = None) -> bytes:
    headers = dict(REQUEST_HEADERS)
    if referer:
        headers["Origin"] = "https://www.invesco.com"
        headers["Referer"] = referer
    request = Request(url, headers=headers)
    with urlopen(request, timeout=120) as response:
        data = response.read()
    if not data:
        raise RuntimeError(f"Provider returned an empty response: {url}")
    return data


def number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[^0-9.()\-]", "", str(value))
    if not cleaned:
        return None
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = f"-{cleaned[1:-1]}"
    try:
        return float(cleaned)
    except ValueError:
        return None


def iso_date(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    for pattern in ("%Y-%m-%d", "%m/%d/%Y", "%d-%b-%Y", "%d-%B-%Y"):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            continue
    return text


def implied_mark(market_value: float | None, par_value: float | None) -> float | None:
    if market_value is None or par_value is None or par_value <= 0:
        return None
    mark = market_value / par_value * 100
    return round(mark, 4) if 0 < mark < 250 else None


def summarize(ticker: str, name: str, as_of: str, source_page: str, source_data: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
    marks = [row["implied_mark"] for row in rows if row.get("implied_mark") is not None]
    return {
        "ticker": ticker,
        "name": name,
        "as_of": as_of,
        "source_page_url": source_page,
        "source_data_url": source_data,
        "row_count": len(rows),
        "marked_row_count": len(marks),
        "total_market_value": round(sum(row.get("market_value") or 0 for row in rows), 2),
        "total_weight_pct": round(sum(row.get("weight_pct") or 0 for row in rows), 6),
        "median_implied_mark": round(sorted(marks)[len(marks) // 2], 4) if marks else None,
        "holdings": rows,
    }


def parse_srln(data: bytes) -> dict[str, Any]:
    from io import BytesIO

    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    title = str(sheet.cell(3, 2).value or "")
    match = re.search(r"As of\s+(.+)$", title, flags=re.I)
    as_of = iso_date(match.group(1).strip()) if match else None
    if not as_of:
        raise RuntimeError("SRLN workbook did not contain a recognizable as-of date")

    headers = [str(sheet.cell(5, column).value or "").strip() for column in range(1, 9)]
    expected = ["Name", "Identifier", "FIGI", "Weight", "Coupon", "Maturity", "Par Value", "Market Value"]
    if headers != expected:
        raise RuntimeError(f"SRLN holdings columns changed: {headers}")

    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=6, max_col=8, values_only=True):
        name = str(values[0] or "").strip()
        weight = number(values[3])
        market_value = number(values[7])
        if not name or weight is None or market_value is None:
            continue
        par_value = number(values[6])
        rows.append({
            "ticker": None,
            "name": name,
            "identifier": str(values[1]).strip() if values[1] else None,
            "figi": str(values[2]).strip() if values[2] else None,
            "cusip": None,
            "security_type": "Senior loan / fund holding",
            "coupon_pct": number(values[4]),
            "maturity_date": iso_date(values[5]),
            "next_call_date": None,
            "rating": None,
            "par_or_shares": par_value,
            "market_value": market_value,
            "weight_pct": weight,
            "currency": "USD",
            "implied_mark": implied_mark(market_value, par_value),
        })
    if not rows:
        raise RuntimeError("SRLN workbook contained no parseable holdings")
    return summarize("SRLN", "State Street Blackstone Senior Loan ETF", as_of, SRLN_PAGE, SRLN_DOWNLOAD, rows)


def parse_bkln(data: bytes) -> dict[str, Any]:
    payload = json.loads(data.decode("utf-8"))
    as_of = iso_date(payload.get("effectiveBusinessDate") or payload.get("effectiveDate"))
    if not as_of:
        raise RuntimeError("BKLN response did not contain an effective business date")
    rows: list[dict[str, Any]] = []
    for item in payload.get("holdings") or []:
        name = html.unescape(str(item.get("issuerName") or "")).strip()
        if not name:
            continue
        par_value = number(item.get("units"))
        market_value = number(item.get("marketValueBase"))
        rows.append({
            "ticker": str(item.get("ticker") or "").strip() or None,
            "name": name,
            "identifier": None,
            "figi": None,
            "cusip": str(item.get("cusip") or "").strip() or None,
            "security_type": str(item.get("securityTypeName") or "").strip() or None,
            "coupon_pct": number(item.get("coupon")),
            "maturity_date": iso_date(item.get("maturityDate")),
            "next_call_date": iso_date(item.get("nextCallDate")),
            "rating": str(item.get("spMoodysRating") or "").strip() or None,
            "par_or_shares": par_value,
            "market_value": market_value,
            "weight_pct": number(item.get("percentageOfTotalNetAssets")),
            "currency": str(item.get("currency") or "").strip() or None,
            "implied_mark": implied_mark(market_value, par_value),
        })
    expected_count = int(payload.get("totalNumberOfHoldings") or 0)
    if not rows or (expected_count and abs(len(rows) - expected_count) > 5):
        raise RuntimeError(f"BKLN holdings count failed validation: parsed {len(rows)}, provider reported {expected_count}")
    return summarize("BKLN", "Invesco Senior Loan ETF", as_of, BKLN_PAGE, BKLN_API, rows)


def load_prior() -> dict[str, Any]:
    if not OUTPUT_PATH.exists():
        return {}
    try:
        return json.loads(OUTPUT_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def write_snapshot(fund: dict[str, Any]) -> tuple[Path, str]:
    snapshot = {
        "ticker": fund["ticker"],
        "name": fund["name"],
        "as_of": fund["as_of"],
        "source_page_url": fund["source_page_url"],
        "source_data_url": fund["source_data_url"],
        "row_count": fund["row_count"],
        "marked_row_count": fund["marked_row_count"],
        "total_market_value": fund["total_market_value"],
        "total_weight_pct": fund["total_weight_pct"],
        "median_implied_mark": fund["median_implied_mark"],
        "holdings": fund["holdings"],
    }
    encoded = (json.dumps(snapshot, indent=2, ensure_ascii=False) + "\n").encode("utf-8")
    digest = hashlib.sha256(encoded).hexdigest()
    target = ARCHIVE_ROOT / fund["as_of"] / f"{fund['ticker'].lower()}.json"
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(encoded)
    return target, digest


def history_rows(prior: dict[str, Any], funds: list[dict[str, Any]], hashes: dict[str, str]) -> list[dict[str, Any]]:
    history = {
        (str(item.get("ticker")), str(item.get("as_of"))): item
        for item in prior.get("history", [])
        if item.get("ticker") and item.get("as_of")
    }
    for fund in funds:
        relative_path = f"data/loan-etf-holdings/{fund['as_of']}/{fund['ticker'].lower()}.json"
        history[(fund["ticker"], fund["as_of"])] = {
            "ticker": fund["ticker"],
            "as_of": fund["as_of"],
            "row_count": fund["row_count"],
            "total_market_value": fund["total_market_value"],
            "total_weight_pct": fund["total_weight_pct"],
            "sha256": hashes[fund["ticker"]],
            "snapshot_path": relative_path,
        }
    return sorted(history.values(), key=lambda row: (row["as_of"], row["ticker"]))[-1460:]


def main() -> None:
    checked_at = utc_now()
    prior = load_prior()
    prior_by_ticker = {fund["ticker"]: fund for fund in prior.get("funds", []) if fund.get("ticker")}
    funds: list[dict[str, Any]] = []
    statuses: list[dict[str, Any]] = []
    errors: list[str] = []
    hashes: dict[str, str] = {}

    adapters = [
        ("SRLN", SRLN_PAGE, lambda: parse_srln(fetch_bytes(SRLN_DOWNLOAD))),
        ("BKLN", BKLN_PAGE, lambda: parse_bkln(fetch_bytes(BKLN_API, referer=BKLN_PAGE))),
    ]
    for ticker, page_url, adapter in adapters:
        try:
            fund = adapter()
            _, digest = write_snapshot(fund)
            hashes[ticker] = digest
            funds.append(fund)
            statuses.append({
                "ticker": ticker,
                "status": "refreshed",
                "as_of": fund["as_of"],
                "records": fund["row_count"],
                "checked_at_utc": checked_at,
                "source_url": page_url,
                "message": "Official daily holdings refreshed and archived.",
            })
        except Exception as exc:
            message = f"{type(exc).__name__}: {exc}"
            errors.append(f"{ticker}: {message}")
            prior_fund = prior_by_ticker.get(ticker)
            if prior_fund:
                funds.append(prior_fund)
            statuses.append({
                "ticker": ticker,
                "status": "stale_fallback" if prior_fund else "error",
                "as_of": prior_fund.get("as_of") if prior_fund else None,
                "records": prior_fund.get("row_count", 0) if prior_fund else 0,
                "checked_at_utc": checked_at,
                "source_url": page_url,
                "message": message,
            })

    if not funds:
        raise RuntimeError("Neither official ETF source produced holdings and no prior snapshot is available")

    for fund in funds:
        if fund["ticker"] not in hashes:
            matching = next(
                (item for item in prior.get("history", []) if item.get("ticker") == fund["ticker"] and item.get("as_of") == fund.get("as_of")),
                None,
            )
            hashes[fund["ticker"]] = matching.get("sha256", "") if matching else ""

    payload = {
        "meta": {
            "generated_at_utc": checked_at,
            "fund_count": len(funds),
            "latest_as_of": max((fund["as_of"] for fund in funds), default=None),
            "snapshot_count": len({(item.get("ticker"), item.get("as_of")) for item in prior.get("history", [])}) + sum(
                1 for fund in funds if not any(item.get("ticker") == fund["ticker"] and item.get("as_of") == fund["as_of"] for item in prior.get("history", []))
            ),
            "schedule": "Weekdays after the US market close; provider business dates are retained verbatim.",
            "methodology": "Official provider rows are normalized without issuer-name matching. Implied mark equals market value divided by par/shares only when both fields are positive and the ratio is plausible.",
            "source_status": statuses,
            "errors": errors,
        },
        "funds": sorted(funds, key=lambda fund: fund["ticker"]),
        "history": history_rows(prior, funds, hashes),
        "limitations": [
            "Provider holdings are not executable bids and may include cash, money-market funds, bonds, equities, or administrative rows in addition to loans.",
            "Market value divided by par/shares is a reference mark, not a promise that the instrument can be traded at that level.",
            "SRLN and BKLN may hold different facilities, currencies, or tranches of the same borrower, so name overlap alone is not a like-for-like valuation comparison.",
        ],
    }
    payload["meta"]["snapshot_count"] = len(payload["history"])
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    print(f"Wrote {OUTPUT_PATH.relative_to(PROJECT_ROOT)}")
    for status in statuses:
        print(f"{status['ticker']}: {status['status']} ({status['records']} rows; as of {status['as_of']})")
    if errors:
        print("Refresh warnings: " + " | ".join(errors))


if __name__ == "__main__":
    main()
