from __future__ import annotations

import argparse
import base64
import csv
import io
import json
import os
import re
import sys
import time
from collections import defaultdict
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import quote, urlencode

sys.path.insert(0, str(Path(__file__).resolve().parent))

from free_sources import (  # noqa: E402
    cached_bytes,
    cached_json,
    decode_delimited,
    find_zip_member,
    iter_zip_dict_rows,
    newest_zip_links,
    normalize_entity,
    number,
    page_links,
    pick,
    request_bytes,
    request_json_post,
    safe_error,
    status_row,
    utc_now,
    zip_members,
    zip_sort_key,
)


PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_PATH = PROJECT_ROOT / "lib" / "free-source-intelligence.json"
DASHBOARD_PATH = PROJECT_ROOT / "lib" / "dashboard-data.json"
FUNDING_PATH = PROJECT_ROOT / "lib" / "bdc-funding-market.json"
SIGNALS_PATH = PROJECT_ROOT / "lib" / "research-signals.json"

FUNDS: dict[str, dict[str, Any]] = {
    "ARCC": {"cik": 1287750, "name": "Ares Capital Corporation"},
    "BBDC": {"cik": 1379785, "name": "Barings BDC Inc"},
    "BXSL": {"cik": 1736035, "name": "Blackstone Secured Lending Fund"},
    "FSK": {"cik": 1422183, "name": "FS KKR Capital Corp"},
    "GBDC": {"cik": 1476765, "name": "Golub Capital BDC Inc"},
    "MAIN": {"cik": 1396440, "name": "Main Street Capital Corporation"},
    "OBDC": {"cik": 1655888, "name": "Blue Owl Capital Corporation"},
    "TSLX": {"cik": 1508655, "name": "Sixth Street Specialty Lending Inc"},
}

SEC_PAGES = {
    "bdc": "https://www.sec.gov/data-research/sec-markets-data/bdc-data-sets",
    "form_d": "https://www.sec.gov/data-research/sec-markets-data/form-d-data-sets",
    "insider": "https://www.sec.gov/data-research/sec-markets-data/insider-transactions-data-sets",
    "13f": "https://www.sec.gov/data-research/sec-markets-data/form-13f-data-sets",
    "notes": "https://www.sec.gov/data-research/sec-markets-data/financial-statement-notes-data-sets",
}

ETF_SOURCES: dict[str, dict[str, Any]] = {
    "SRLN": {
        "name": "SPDR Blackstone Senior Loan ETF",
        "page": "https://www.ssga.com/us/en/individual/etfs/state-street-blackstone-senior-loan-etf-srln",
        "candidates": [
            "https://www.ssga.com/library-content/products/fund-data/etfs/us/holdings-daily-us-en-srln.xlsx",
        ],
    },
    "BKLN": {
        "name": "Invesco Senior Loan ETF",
        "page": "https://www.invesco.com/us/financial-products/etfs/product-detail?audienceType=Investor&ticker=BKLN",
        "candidates": [
            "https://www.invesco.com/us/financial-products/etfs/holdings/main/holdings/0?action=download&audienceType=Institutional&ticker=BKLN",
        ],
    },
    "FLBL": {
        "name": "Franklin Senior Loan ETF",
        "page": "https://www.franklintempleton.com/investments/options/exchange-traded-funds/products/29064/SINGLCLASS/franklin-senior-loan-etf/FLBL",
        "candidates": [],
    },
}

FRED_SERIES = {
    "BAMLH0A0HYM2": "US high-yield option-adjusted spread",
    "BAMLH0A1HYBB": "BB high-yield option-adjusted spread",
    "BAMLH0A2HYB": "Single-B high-yield option-adjusted spread",
    "BAMLH0A3HYC": "CCC and lower option-adjusted spread",
    "DRTSCILM": "Banks tightening C&I standards for large and middle-market firms",
    "DRTSCIS": "Banks tightening C&I standards for small firms",
}


def period_label(row: dict[str, str]) -> str:
    text = f"{row.get('label', '')} {row.get('url', '')}"
    year_match = re.search(r"(20\d{2})", text)
    quarter_match = re.search(r"(?:q|quarter\D*)([1-4])", text, re.I)
    month_match = re.search(r"20\d{2}[_-](0?[1-9]|1[0-2])", text)
    if year_match and month_match:
        return f"{year_match.group(1)}-{int(month_match.group(1)):02d}"
    if year_match and quarter_match:
        return f"{year_match.group(1)} Q{quarter_match.group(1)}"
    return row.get("label") or Path(row.get("url", "")).stem


def sec_archive_url(cik: int | str, accession: str) -> str:
    compact = re.sub(r"\D", "", accession)
    return f"https://www.sec.gov/Archives/edgar/data/{int(cik)}/{compact}/"


def borrower_universe(dashboard: dict[str, Any]) -> list[dict[str, Any]]:
    rows = []
    for row in dashboard.get("cross_fund_issuer_latest", []):
        name = str(row.get("representative_issuer_name") or row.get("issuer_match_key") or "").strip()
        normalized = normalize_entity(name)
        if not normalized:
            continue
        rows.append({
            "issuer_match_key": row.get("issuer_match_key"),
            "name": name,
            "normalized": normalized,
            "funds": row.get("funds", []),
            "fair_value_mm": float(row.get("fair_value_mm") or 0),
        })
    rows.sort(key=lambda item: (-item["fair_value_mm"], item["name"]))
    return rows


def load_sec_zip(page_url: str, *, candidates: Iterable[str], fallback_limit: int = 1) -> tuple[dict[str, str], bytes, str]:
    links, _ = newest_zip_links(page_url, limit=max(fallback_limit, 1))
    if not links:
        raise RuntimeError("No ZIP links were found on the SEC dataset page")
    last_message = ""
    for link in links:
        response = cached_bytes(link["url"], suffix=".zip", max_age_hours=168, timeout=180)
        member = find_zip_member(response.data, candidates)
        if member is not None:
            return link, response.data, response.fetched_at_utc
        last_message = f"{period_label(link)} did not contain {', '.join(candidates)}"
    raise RuntimeError(last_message or "Expected table not found in SEC ZIP")


def ingest_sec_bdc() -> tuple[dict[str, Any], dict[str, Any]]:
    page = SEC_PAGES["bdc"]
    links, _ = newest_zip_links(page, limit=8)
    if not links:
        raise RuntimeError("No monthly BDC packages found")
    selected: tuple[dict[str, str], bytes, str] | None = None
    checked: list[str] = []
    for link in links:
        response = cached_bytes(link["url"], suffix=".zip", max_age_hours=168, timeout=240)
        checked.append(period_label(link))
        soi_member = find_zip_member(response.data, ("soi.tsv", "soi.txt"))
        if soi_member is None:
            continue
        soi_count = sum(1 for _ in iter_zip_dict_rows(response.data, ("soi.tsv", "soi.txt")))
        if soi_count:
            selected = (link, response.data, response.fetched_at_utc)
            break
    if selected is None:
        raise RuntimeError(f"No populated Schedule of Investments table in checked packages: {', '.join(checked)}")
    link, zip_data, fetched_at = selected
    target_ciks = {str(config["cik"]): ticker for ticker, config in FUNDS.items()}
    submissions: dict[str, dict[str, Any]] = {}
    for row in iter_zip_dict_rows(zip_data, ("sub.tsv", "sub.txt")):
        cik_text = str(pick(row, "cik") or "").lstrip("0")
        if cik_text not in target_ciks:
            continue
        accession = str(pick(row, "adsh", "accession_number") or "")
        submissions[accession] = {
            "ticker": target_ciks[cik_text],
            "cik": int(cik_text),
            "form": pick(row, "form"),
            "filed": pick(row, "filed"),
            "period": pick(row, "period"),
        }

    schedule_rows_by_fund: dict[str, int] = defaultdict(int)
    schedule_fair_value_by_fund: dict[str, float] = defaultdict(float)
    for row in iter_zip_dict_rows(zip_data, ("soi.tsv", "soi.txt")):
        submission = submissions.get(str(pick(row, "adsh", "accession_number") or ""))
        if submission is None:
            continue
        ticker = submission["ticker"]
        schedule_rows_by_fund[ticker] += 1
        value = number(pick(row, "value", "fair_value", "amount"))
        if value is not None:
            schedule_fair_value_by_fund[ticker] += value

    debt_pattern = re.compile(r"debt|borrow|creditfacility|revolver|seniornote|securednote|unsecurednote|repurchase|redemption|tender", re.I)
    debt_candidates: list[dict[str, Any]] = []
    numeric_count = 0
    for row in iter_zip_dict_rows(zip_data, ("num.tsv", "num.txt")):
        accession = str(pick(row, "adsh", "accession_number") or "")
        submission = submissions.get(accession)
        if submission is None:
            continue
        numeric_count += 1
        tag = str(pick(row, "tag", "element") or "")
        if not debt_pattern.search(tag):
            continue
        debt_candidates.append({
            "ticker": submission["ticker"],
            "filing_period": submission.get("period"),
            "filed": submission.get("filed"),
            "form": submission.get("form"),
            "tag": tag,
            "value": number(pick(row, "value")),
            "unit": pick(row, "uom", "unit"),
            "fact_date": pick(row, "ddate", "date"),
            "quarters": pick(row, "qtrs", "quarters"),
            "accession": accession,
            "source_url": sec_archive_url(submission["cik"], accession),
            "confidence": "raw-candidate",
        })
    debt_candidates.sort(key=lambda row: (str(row["filed"]), row["ticker"], row["tag"]), reverse=True)
    debt_candidates = debt_candidates[:250]

    text_count = 0
    text_candidates: list[dict[str, Any]] = []
    for row in iter_zip_dict_rows(zip_data, ("txt.tsv", "txt.txt")):
        accession = str(pick(row, "adsh", "accession_number") or "")
        submission = submissions.get(accession)
        if submission is None:
            continue
        text_count += 1
        tag = str(pick(row, "tag", "element") or "")
        if debt_pattern.search(tag):
            text_candidates.append({
                "ticker": submission["ticker"],
                "filed": submission.get("filed"),
                "tag": tag,
                "fact_date": pick(row, "ddate", "date"),
                "accession": accession,
                "source_url": sec_archive_url(submission["cik"], accession),
                "confidence": "raw-candidate",
            })
    text_candidates = text_candidates[:100]

    payload = {
        "package_period": period_label(link),
        "package_url": link["url"],
        "fetched_at_utc": fetched_at,
        "checked_periods": checked,
        "members": zip_members(zip_data),
        "filings": list(submissions.values()),
        "schedule_rows_by_fund": dict(sorted(schedule_rows_by_fund.items())),
        "schedule_fair_value_raw_by_fund": dict(sorted(schedule_fair_value_by_fund.items())),
        "numeric_fact_rows": numeric_count,
        "text_fact_rows": text_count,
        "debt_numeric_candidates": debt_candidates,
        "debt_text_candidates": text_candidates,
        "notes_dataset_page": SEC_PAGES["notes"],
    }
    status = status_row(
        "sec_bdc_monthly",
        "SEC monthly BDC data sets",
        "SEC filings",
        "refreshed",
        page,
        records=sum(schedule_rows_by_fund.values()),
        as_of=payload["package_period"],
        message="Newest populated monthly package; raw debt-note facts remain reconciliation candidates.",
    )
    return payload, status


def ingest_form_d(borrowers: list[dict[str, Any]]) -> tuple[dict[str, Any], dict[str, Any]]:
    link, zip_data, fetched_at = load_sec_zip(SEC_PAGES["form_d"], candidates=("issuer",), fallback_limit=2)
    borrower_map: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for borrower in borrowers:
        borrower_map[borrower["normalized"]].append(borrower)
    matches: list[dict[str, Any]] = []
    matched_accessions: set[str] = set()
    for row in iter_zip_dict_rows(zip_data, ("issuer",)):
        issuer_name = str(pick(row, "issuername", "issuer_name") or "").strip()
        normalized = normalize_entity(issuer_name)
        if normalized not in borrower_map:
            continue
        accession = str(pick(row, "accessionnumber", "accession_number") or "")
        issuer_cik = str(pick(row, "issuercik", "issuer_cik", "cik") or "").lstrip("0")
        matched_accessions.add(accession)
        for borrower in borrower_map[normalized]:
            matches.append({
                "issuer_match_key": borrower["issuer_match_key"],
                "dashboard_name": borrower["name"],
                "form_d_issuer_name": issuer_name,
                "issuer_cik": int(issuer_cik) if issuer_cik.isdigit() else None,
                "accession": accession,
                "entity_type": pick(row, "entitytype", "entity_type"),
                "jurisdiction": pick(row, "jurisdictionofinc", "jurisdiction_of_inc"),
                "year_of_incorporation": pick(row, "yearofinc", "year_of_inc"),
                "source_url": sec_archive_url(issuer_cik, accession) if issuer_cik.isdigit() else SEC_PAGES["form_d"],
                "match_method": "exact-normalized-name",
                "confidence": "high-candidate",
            })
    offering_by_accession: dict[str, dict[str, Any]] = {}
    if matched_accessions:
        for row in iter_zip_dict_rows(zip_data, ("offering",)):
            accession = str(pick(row, "accessionnumber", "accession_number") or "")
            if accession not in matched_accessions:
                continue
            offering_by_accession[accession] = {
                "industry_group": pick(row, "industrygroup", "industry_group"),
                "date_of_first_sale": pick(row, "dateoffirstsale", "date_of_first_sale"),
                "total_offering_amount": number(pick(row, "totalofferingamount", "total_offering_amount")),
                "total_amount_sold": number(pick(row, "totalamountsold", "total_amount_sold")),
            }
    for match in matches:
        match["offering"] = offering_by_accession.get(match["accession"])
    matches.sort(key=lambda row: (row["dashboard_name"], row["accession"]), reverse=False)
    payload = {
        "package_period": period_label(link),
        "package_url": link["url"],
        "fetched_at_utc": fetched_at,
        "matches": matches[:500],
        "methodology": "Exact matches after punctuation and common legal-suffix normalization. Matches are candidates and never overwrite dashboard issuer keys automatically.",
    }
    status = status_row(
        "sec_form_d",
        "SEC Form D bulk data",
        "Entity resolution",
        "refreshed",
        SEC_PAGES["form_d"],
        records=len(matches),
        as_of=payload["package_period"],
        confidence="high-candidate",
        message="Exact normalized-name candidates only; manual confirmation is still required.",
    )
    return payload, status


def parse_table_bytes(data: bytes) -> tuple[list[str], list[list[Any]]]:
    if data[:2] == b"PK":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl is required for ETF XLSX holdings") from exc
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        worksheet = workbook.active
        rows = [[cell for cell in row] for row in worksheet.iter_rows(values_only=True)]
    else:
        text = decode_delimited(data)
        dialect = csv.excel_tab if text.count("\t") > text.count(",") else csv.excel
        rows = [list(row) for row in csv.reader(io.StringIO(text), dialect=dialect)]
    header_index = -1
    for index, row in enumerate(rows[:80]):
        normalized = [re.sub(r"[^a-z0-9]", "", str(value or "").lower()) for value in row]
        joined = " ".join(normalized)
        if any(value in {"name", "securityname", "holdingname", "issuername", "description"} for value in normalized) and (
            "marketvalue" in joined or "notional" in joined or "parvalue" in joined or "shares" in joined
        ):
            header_index = index
            break
    if header_index < 0:
        raise RuntimeError("Could not locate a holdings header row")
    headers = [str(value or "").strip() for value in rows[header_index]]
    return headers, rows[header_index + 1 :]


def parse_etf_holdings(data: bytes, ticker: str, borrowers: list[dict[str, Any]], source_url: str) -> dict[str, Any]:
    headers, raw_rows = parse_table_bytes(data)
    normalized_headers = [re.sub(r"[^a-z0-9]", "", header.lower()) for header in headers]

    def column(*candidates: str) -> int | None:
        for candidate in candidates:
            normalized_candidate = re.sub(r"[^a-z0-9]", "", candidate.lower())
            for index, header in enumerate(normalized_headers):
                if header == normalized_candidate or normalized_candidate in header:
                    return index
        return None

    name_index = column("security name", "holding name", "issuer name", "description", "name")
    market_index = column("market value", "notional market value")
    par_index = column("par value", "par amount", "face value", "principal", "shares/par value", "shares")
    cusip_index = column("cusip")
    maturity_index = column("maturity")
    asset_index = column("asset class", "security type", "type")
    date_index = column("as of date", "date")
    if name_index is None:
        raise RuntimeError("ETF holdings file has no recognizable security-name column")
    borrower_map: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for borrower in borrowers:
        borrower_map[borrower["normalized"]].append(borrower)
    holdings: list[dict[str, Any]] = []
    as_of_values: set[str] = set()
    for values in raw_rows:
        if name_index >= len(values):
            continue
        name = str(values[name_index] or "").strip()
        if not name or name.lower() in {"total", "cash", "cash and equivalents"}:
            continue
        market_value = number(values[market_index]) if market_index is not None and market_index < len(values) else None
        par_value = number(values[par_index]) if par_index is not None and par_index < len(values) else None
        implied_mark = market_value / par_value * 100 if market_value is not None and par_value and par_value > 0 else None
        if implied_mark is not None and not (0 < implied_mark < 250):
            implied_mark = None
        normalized = normalize_entity(name)
        matches = borrower_map.get(normalized, [])
        row_date = str(values[date_index] or "").strip() if date_index is not None and date_index < len(values) else ""
        if row_date:
            as_of_values.add(row_date)
        holdings.append({
            "etf": ticker,
            "name": name,
            "cusip": str(values[cusip_index] or "").strip() if cusip_index is not None and cusip_index < len(values) else None,
            "maturity": str(values[maturity_index] or "").strip() if maturity_index is not None and maturity_index < len(values) else None,
            "asset_class": str(values[asset_index] or "").strip() if asset_index is not None and asset_index < len(values) else None,
            "market_value": market_value,
            "par_or_shares": par_value,
            "implied_mark": round(implied_mark, 6) if implied_mark is not None else None,
            "issuer_match_keys": [match["issuer_match_key"] for match in matches],
            "match_method": "exact-normalized-name" if matches else None,
            "confidence": "source-direct" if implied_mark is not None else "holding-only",
            "source_url": source_url,
        })
    marked = [row for row in holdings if row["implied_mark"] is not None]
    matched = [row for row in holdings if row["issuer_match_keys"]]
    return {
        "etf": ticker,
        "as_of": sorted(as_of_values)[-1] if as_of_values else None,
        "row_count": len(holdings),
        "marked_row_count": len(marked),
        "matched_bdc_borrower_count": len(matched),
        "source_url": source_url,
        "holdings": holdings[:2500],
    }


def ingest_etf_holdings(borrowers: list[dict[str, Any]]) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    funds: list[dict[str, Any]] = []
    statuses: list[dict[str, Any]] = []
    for ticker, config in ETF_SOURCES.items():
        candidate_urls = list(config["candidates"])
        try:
            links, _ = page_links(config["page"], max_age_hours=24)
            discovered = [
                url for url, label in links
                if re.search(r"holdings|download", label, re.I) and re.search(r"\.xlsx?|\.csv(?:\?|$)", url, re.I)
            ]
            candidate_urls = candidate_urls + [url for url in discovered if url not in candidate_urls]
        except Exception:
            pass
        last_error = "No public holdings-download URL was found"
        parsed: dict[str, Any] | None = None
        for url in candidate_urls:
            try:
                response = cached_bytes(url, max_age_hours=12, timeout=120)
                prefix = response.data[:1500].lower()
                if b"<!doctype html" in prefix or b"<html" in prefix:
                    raise RuntimeError("Provider returned an HTML page instead of a holdings file")
                parsed = parse_etf_holdings(response.data, ticker, borrowers, url)
                parsed["fetched_at_utc"] = response.fetched_at_utc
                break
            except Exception as exc:
                last_error = safe_error(exc)
        if parsed is None:
            funds.append({"etf": ticker, "as_of": None, "row_count": 0, "holdings": [], "error": last_error, "source_url": config["page"]})
            statuses.append(status_row(
                f"etf_{ticker.lower()}", config["name"], "Daily loan marks", "available_not_refreshed", config["page"],
                message=last_error,
            ))
        else:
            funds.append(parsed)
            statuses.append(status_row(
                f"etf_{ticker.lower()}", config["name"], "Daily loan marks", "refreshed", config["page"],
                records=parsed["row_count"], as_of=parsed["as_of"],
                message=f"{parsed['marked_row_count']} rows support an inferred market-value/par mark; {parsed['matched_bdc_borrower_count']} exact borrower matches.",
            ))
    return {"funds": funds, "methodology": "Daily issuer files are preserved as source rows. An implied mark is calculated only when both market value and par/shares are present and the ratio is plausible."}, statuses


def ingest_gleif(borrowers: list[dict[str, Any]]) -> tuple[dict[str, Any], dict[str, Any]]:
    limit = max(0, int(os.environ.get("BDC_GLEIF_QUERY_LIMIT", "12")))
    candidates: list[dict[str, Any]] = []
    errors: list[str] = []
    for borrower in borrowers[:limit]:
        url = "https://api.gleif.org/api/v1/lei-records?" + urlencode({"filter[entity.legalName]": borrower["name"], "page[size]": 5})
        try:
            response, _ = cached_json(url, max_age_hours=168)
            for item in response.get("data", []):
                attributes = item.get("attributes", {})
                entity = attributes.get("entity", {})
                legal_name = entity.get("legalName", {}).get("name")
                if not legal_name:
                    continue
                exact = normalize_entity(legal_name) == borrower["normalized"]
                candidates.append({
                    "issuer_match_key": borrower["issuer_match_key"],
                    "dashboard_name": borrower["name"],
                    "legal_name": legal_name,
                    "lei": attributes.get("lei") or item.get("id"),
                    "entity_status": entity.get("status"),
                    "jurisdiction": entity.get("jurisdiction"),
                    "headquarters": entity.get("headquartersAddress"),
                    "match_method": "exact-normalized-name" if exact else "provider-search-candidate",
                    "confidence": "high-candidate" if exact else "review-candidate",
                    "source_url": f"https://api.gleif.org/api/v1/lei-records/{item.get('id')}",
                })
        except Exception as exc:
            errors.append(f"{borrower['name']}: {safe_error(exc)}")
        time.sleep(0.08)
    status = status_row(
        "gleif",
        "GLEIF LEI API",
        "Entity resolution",
        "refreshed" if candidates or not errors else "available_not_refreshed",
        "https://www.gleif.org/en/lei-data/gleif-api",
        records=len(candidates),
        confidence="candidate",
        message=f"Queried {limit} highest-exposure borrowers. " + (f"{len(errors)} query errors." if errors else "Results remain review candidates."),
    )
    return {"query_limit": limit, "candidates": candidates, "errors": errors}, status


def ingest_company_registries(borrowers: list[dict[str, Any]]) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    output: dict[str, Any] = {"companies_house": [], "opencorporates": []}
    statuses: list[dict[str, Any]] = []
    query_rows = borrowers[: max(0, int(os.environ.get("BDC_REGISTRY_QUERY_LIMIT", "8")))]
    companies_house_key = os.environ.get("COMPANIES_HOUSE_API_KEY")
    if not companies_house_key:
        statuses.append(status_row(
            "companies_house", "Companies House API", "Entity resolution", "credential_required",
            "https://developer.company-information.service.gov.uk/", credential_env="COMPANIES_HOUSE_API_KEY",
            message="Connector is implemented; set the free API key to query UK company candidates.",
        ))
    else:
        auth = base64.b64encode(f"{companies_house_key}:".encode()).decode()
        for borrower in query_rows:
            url = "https://api.company-information.service.gov.uk/search/companies?" + urlencode({"q": borrower["name"], "items_per_page": 5})
            try:
                response, _ = cached_json(url, max_age_hours=168, headers={"Authorization": f"Basic {auth}"})
                for item in response.get("items", []):
                    output["companies_house"].append({
                        "issuer_match_key": borrower["issuer_match_key"], "dashboard_name": borrower["name"],
                        "company_name": item.get("title"), "company_number": item.get("company_number"),
                        "company_status": item.get("company_status"), "date_of_creation": item.get("date_of_creation"),
                        "address": item.get("address"), "confidence": "review-candidate",
                        "source_url": f"https://find-and-update.company-information.service.gov.uk/company/{item.get('company_number')}",
                    })
            except Exception:
                continue
        statuses.append(status_row(
            "companies_house", "Companies House API", "Entity resolution", "refreshed",
            "https://developer.company-information.service.gov.uk/", records=len(output["companies_house"]), confidence="candidate",
            message="Provider search candidates; no automatic issuer-key mutation.",
        ))
    opencorporates_token = os.environ.get("OPENCORPORATES_API_TOKEN")
    if not opencorporates_token:
        statuses.append(status_row(
            "opencorporates", "OpenCorporates API", "Entity resolution", "credential_required",
            "https://api.opencorporates.com/documentation/API-Reference", credential_env="OPENCORPORATES_API_TOKEN",
            message="Connector is implemented; an API token is required by the current service.",
        ))
    else:
        for borrower in query_rows:
            url = "https://api.opencorporates.com/v0.4/companies/search?" + urlencode({"q": borrower["name"], "api_token": opencorporates_token, "per_page": 5})
            try:
                response = json.loads(request_bytes(url, timeout=90).decode("utf-8"))
                companies = response.get("results", {}).get("companies", [])
                for wrapped in companies:
                    item = wrapped.get("company", {})
                    output["opencorporates"].append({
                        "issuer_match_key": borrower["issuer_match_key"], "dashboard_name": borrower["name"],
                        "company_name": item.get("name"), "company_number": item.get("company_number"),
                        "jurisdiction_code": item.get("jurisdiction_code"), "current_status": item.get("current_status"),
                        "incorporation_date": item.get("incorporation_date"), "confidence": "review-candidate",
                        "source_url": item.get("opencorporates_url"),
                    })
            except Exception:
                continue
        statuses.append(status_row(
            "opencorporates", "OpenCorporates API", "Entity resolution", "refreshed",
            "https://api.opencorporates.com/documentation/API-Reference", records=len(output["opencorporates"]), confidence="candidate",
            message="Provider search candidates; licensing terms still apply to reuse.",
        ))
    return output, statuses


def ingest_openfigi(funding: dict[str, Any]) -> tuple[dict[str, Any], dict[str, Any]]:
    cusips: list[str] = []
    for row in funding.get("series", []):
        cusip = re.sub(r"\W", "", str(row.get("cusip") or "")).upper()
        if len(cusip) == 9 and cusip not in cusips:
            cusips.append(cusip)
    jobs = [{"idType": "ID_CUSIP", "idValue": cusip} for cusip in cusips[:100]]
    results: list[dict[str, Any]] = []
    api_key = os.environ.get("OPENFIGI_API_KEY")
    headers = {"X-OPENFIGI-APIKEY": api_key} if api_key else {}
    batch_size = 100 if api_key else 10
    for offset in range(0, len(jobs), batch_size):
        batch = jobs[offset : offset + batch_size]
        try:
            response = request_json_post("https://api.openfigi.com/v3/mapping", batch, headers=headers)
            for job, item in zip(batch, response):
                results.append({"cusip": job["idValue"], "data": item.get("data", []), "error": item.get("error")})
        except Exception as exc:
            for job in batch:
                results.append({"cusip": job["idValue"], "data": [], "error": safe_error(exc)})
        time.sleep(0.15)
    matched = sum(1 for row in results if row["data"])
    status = status_row(
        "openfigi", "OpenFIGI mapping API", "Instrument identity", "refreshed" if results else "available_not_refreshed",
        "https://www.openfigi.com/api", records=matched,
        message=f"Mapped {matched} of {len(results)} BDC note CUSIPs; API key is optional and only increases rate limits.",
        credential_env="OPENFIGI_API_KEY (optional)",
    )
    return {"queries": results}, status


def ingest_massive_market() -> tuple[dict[str, Any], dict[str, Any]]:
    api_key = os.environ.get("MASSIVE_API_KEY")
    source_url = "https://massive.com/docs/rest/stocks/overview"
    if not api_key:
        return {"quotes": [], "dividends": [], "splits": []}, status_row(
            "massive", "Massive market data API", "Market data", "credential_required", source_url,
            credential_env="MASSIVE_API_KEY", message="Connector is implemented; no market snapshot was replaced without a configured key.",
        )
    base = os.environ.get("MASSIVE_API_BASE", "https://api.massive.com").rstrip("/")
    headers = {"Authorization": f"Bearer {api_key}"}
    quotes: list[dict[str, Any]] = []
    dividends: list[dict[str, Any]] = []
    splits: list[dict[str, Any]] = []
    errors: list[str] = []
    for ticker in FUNDS:
        try:
            payload, _ = cached_json(f"{base}/v2/aggs/ticker/{ticker}/prev?adjusted=true", max_age_hours=6, headers=headers)
            results = payload.get("results") or []
            if results:
                item = results[0]
                timestamp = item.get("t")
                price_date = datetime.fromtimestamp(timestamp / 1000, timezone.utc).date().isoformat() if timestamp else None
                quotes.append({
                    "ticker": ticker, "price": item.get("c"), "open": item.get("o"), "high": item.get("h"),
                    "low": item.get("l"), "volume": item.get("v"), "price_date": price_date,
                    "source_url": f"{source_url}#previous-close", "confidence": "source-direct",
                })
            dividend_payload, _ = cached_json(
                f"{base}/v3/reference/dividends?{urlencode({'ticker': ticker, 'limit': 5, 'sort': 'ex_dividend_date', 'order': 'desc'})}",
                max_age_hours=24, headers=headers,
            )
            for item in dividend_payload.get("results", []):
                dividends.append({"ticker": ticker, **item, "source_url": source_url})
            split_payload, _ = cached_json(
                f"{base}/v3/reference/splits?{urlencode({'ticker': ticker, 'limit': 5, 'sort': 'execution_date', 'order': 'desc'})}",
                max_age_hours=24, headers=headers,
            )
            for item in split_payload.get("results", []):
                splits.append({"ticker": ticker, **item, "source_url": source_url})
        except Exception as exc:
            errors.append(f"{ticker}: {safe_error(exc)}")
    status = status_row(
        "massive", "Massive market data API", "Market data", "refreshed" if quotes else "available_not_refreshed", source_url,
        records=len(quotes), as_of=max((row["price_date"] for row in quotes if row["price_date"]), default=None),
        message=f"{len(dividends)} dividend and {len(splits)} split observations. " + (f"{len(errors)} ticker errors." if errors else ""),
    )
    return {"quotes": quotes, "dividends": dividends, "splits": splits, "errors": errors}, status


def ingest_fred() -> tuple[dict[str, Any], dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    errors: list[str] = []
    for series_id, label in FRED_SERIES.items():
        url = f"https://fred.stlouisfed.org/graph/fredgraph.csv?id={series_id}"
        try:
            response = cached_bytes(url, suffix=".csv", max_age_hours=24)
            observations = list(csv.DictReader(io.StringIO(response.data.decode("utf-8-sig"))))
            value_key = next((key for key in (observations[0].keys() if observations else []) if key != "observation_date"), series_id)
            latest = next((item for item in reversed(observations) if number(item.get(value_key)) is not None), None)
            if latest:
                rows.append({
                    "series_id": series_id, "label": label, "observation_date": latest.get("observation_date"),
                    "value": number(latest.get(value_key)), "source_url": f"https://fred.stlouisfed.org/series/{series_id}",
                    "confidence": "source-direct",
                })
        except Exception as exc:
            errors.append(f"{series_id}: {safe_error(exc)}")
    status = status_row(
        "fred", "Federal Reserve Economic Data (FRED)", "Macro drivers", "refreshed" if rows else "available_not_refreshed",
        "https://fred.stlouisfed.org/", records=len(rows), as_of=max((row["observation_date"] for row in rows), default=None),
        message="Public CSV observations for high-yield spreads and SLOOS lending standards. " + (f"{len(errors)} series errors." if errors else ""),
    )
    return {"series": rows, "errors": errors}, status


def ingest_insiders() -> tuple[dict[str, Any], dict[str, Any]]:
    link, zip_data, fetched_at = load_sec_zip(SEC_PAGES["insider"], candidates=("submission",), fallback_limit=2)
    target_ciks = {str(config["cik"]): ticker for ticker, config in FUNDS.items()}
    submissions: dict[str, dict[str, Any]] = {}
    for row in iter_zip_dict_rows(zip_data, ("submission",)):
        cik = str(pick(row, "issuercik", "issuer_cik", "cik") or "").lstrip("0")
        if cik not in target_ciks:
            continue
        accession = str(pick(row, "accessionnumber", "accession_number") or "")
        submissions[accession] = {
            "ticker": target_ciks[cik], "cik": int(cik), "accession": accession,
            "filing_date": pick(row, "filingdate", "filing_date"), "period_of_report": pick(row, "periodofreport", "period_of_report"),
        }
    owners: dict[str, list[str]] = defaultdict(list)
    for row in iter_zip_dict_rows(zip_data, ("reportingowner", "reporting_owner", "owner")):
        accession = str(pick(row, "accessionnumber", "accession_number") or "")
        if accession not in submissions:
            continue
        name = str(pick(row, "rptownername", "reportingownername", "ownername") or "").strip()
        if name:
            owners[accession].append(name)
    transactions: list[dict[str, Any]] = []
    for row in iter_zip_dict_rows(zip_data, ("nonderiv_trans", "nonderivtrans", "non_derivative_transaction")):
        accession = str(pick(row, "accessionnumber", "accession_number") or "")
        submission = submissions.get(accession)
        if submission is None:
            continue
        code = str(pick(row, "transcode", "transactioncode", "transaction_code") or "").upper()
        if code not in {"P", "S"}:
            continue
        transactions.append({
            **submission,
            "owners": owners.get(accession, []),
            "transaction_date": pick(row, "transdate", "transactiondate", "transaction_date"),
            "transaction_code": code,
            "acquired_disposed": pick(row, "transacquireddispcd", "acquireddisposedcode"),
            "shares": number(pick(row, "transshares", "transactionshares", "transaction_shares")),
            "price_per_share": number(pick(row, "transpricepershare", "transactionpricepershare", "transaction_price_per_share")),
            "source_url": sec_archive_url(submission["cik"], accession),
            "confidence": "source-direct",
        })
    transactions.sort(key=lambda row: str(row.get("transaction_date") or row.get("filing_date") or ""), reverse=True)
    status = status_row(
        "sec_insiders", "SEC insider-transactions bulk data", "Ownership and governance", "refreshed", SEC_PAGES["insider"],
        records=len(transactions), as_of=period_label(link), message="Open-market purchase and sale codes (P/S) for the eight covered BDC issuers.",
    )
    return {"package_period": period_label(link), "package_url": link["url"], "fetched_at_utc": fetched_at, "transactions": transactions[:500]}, status


def ingest_13f(enabled: bool) -> tuple[dict[str, Any], dict[str, Any]]:
    if not enabled:
        return {"positions": [], "enabled": False}, status_row(
            "sec_13f", "SEC Form 13F bulk data", "Institutional ownership", "available_not_refreshed", SEC_PAGES["13f"],
            message="Fully implemented but not downloaded by default because the quarterly information table is large. Run with --heavy-13f or BDC_ENABLE_HEAVY_13F=1.",
        )
    link, zip_data, fetched_at = load_sec_zip(SEC_PAGES["13f"], candidates=("infotable", "info_table"), fallback_limit=2)
    aliases: dict[str, str] = {}
    for ticker, config in FUNDS.items():
        aliases[normalize_entity(config["name"])] = ticker
    aggregations: dict[tuple[str, str], dict[str, Any]] = {}
    for row in iter_zip_dict_rows(zip_data, ("infotable", "info_table")):
        issuer = str(pick(row, "nameofissuer", "name_of_issuer") or "")
        ticker = aliases.get(normalize_entity(issuer))
        if ticker is None:
            continue
        accession = str(pick(row, "accessionnumber", "accession_number") or "")
        key = ticker, accession
        aggregate = aggregations.setdefault(key, {"ticker": ticker, "accession": accession, "manager_rows": 0, "value_thousands": 0.0, "shares": 0.0})
        aggregate["manager_rows"] += 1
        aggregate["value_thousands"] += number(pick(row, "value")) or 0
        aggregate["shares"] += number(pick(row, "sshprnamt", "shares")) or 0
    positions = sorted(aggregations.values(), key=lambda row: (-row["value_thousands"], row["ticker"]))
    status = status_row(
        "sec_13f", "SEC Form 13F bulk data", "Institutional ownership", "refreshed", SEC_PAGES["13f"],
        records=len(positions), as_of=period_label(link), message="Issuer-name matched ownership rows; aggregation remains separate from the core valuation score.",
    )
    return {"enabled": True, "package_period": period_label(link), "package_url": link["url"], "fetched_at_utc": fetched_at, "positions": positions}, status


def ingest_courtlistener(signals: dict[str, Any]) -> tuple[dict[str, Any], dict[str, Any]]:
    token = os.environ.get("COURTLISTENER_API_TOKEN")
    source_url = "https://www.courtlistener.com/help/api/rest/"
    if not token:
        return {"alerts": []}, status_row(
            "courtlistener", "CourtListener search API", "Legal events", "credential_required", source_url,
            credential_env="COURTLISTENER_API_TOKEN", message="Connector is implemented. A free account/token is required; no absence-of-litigation inference is made.",
        )
    query_limit = max(0, int(os.environ.get("BDC_LEGAL_QUERY_LIMIT", "10")))
    priority = sorted(signals.get("issuer_signals", []), key=lambda row: int(row.get("priority_rank") or 999999))[:query_limit]
    alerts: list[dict[str, Any]] = []
    errors: list[str] = []
    headers = {"Authorization": f"Token {token}"}
    for issuer in priority:
        name = str(issuer.get("mapped_company") or issuer.get("display_name") or "").strip()
        if not name:
            continue
        url = "https://www.courtlistener.com/api/rest/v3/search/?" + urlencode({"q": f'"{name}"', "type": "r", "order_by": "dateFiled desc"})
        try:
            payload, _ = cached_json(url, max_age_hours=12, headers=headers)
            for item in payload.get("results", [])[:10]:
                alerts.append({
                    "issuer_match_key": issuer.get("issuer_match_key"), "dashboard_name": name,
                    "case_name": item.get("caseName") or item.get("case_name"), "court": item.get("court_citation_string") or item.get("court"),
                    "date_filed": item.get("dateFiled") or item.get("date_filed"), "docket_number": item.get("docketNumber") or item.get("docket_number"),
                    "absolute_url": "https://www.courtlistener.com" + item.get("absolute_url", "") if item.get("absolute_url") else source_url,
                    "match_method": "quoted-name-search", "confidence": "review-candidate",
                })
        except Exception as exc:
            errors.append(f"{name}: {safe_error(exc)}")
    status = status_row(
        "courtlistener", "CourtListener search API", "Legal events", "refreshed" if alerts or not errors else "available_not_refreshed", source_url,
        records=len(alerts), confidence="review-candidate",
        message=f"Queried {query_limit} highest-priority issuers; every result requires party-name and docket review. " + (f"{len(errors)} errors." if errors else ""),
    )
    return {"query_limit": query_limit, "alerts": alerts, "errors": errors}, status


def run_component(
    component_id: str,
    name: str,
    category: str,
    url: str,
    function: Any,
    *args: Any,
) -> tuple[Any, list[dict[str, Any]]]:
    try:
        payload, statuses = function(*args)
        return payload, statuses if isinstance(statuses, list) else [statuses]
    except Exception as exc:
        return {}, [status_row(
            component_id, name, category, "error", url, confidence="unavailable",
            message=safe_error(exc),
        )]


def main() -> None:
    parser = argparse.ArgumentParser(description="Refresh free external intelligence sources used by the BDC dashboard.")
    parser.add_argument("--heavy-13f", action="store_true", help="Download and parse the large quarterly SEC 13F information table.")
    args = parser.parse_args()
    dashboard = json.loads(DASHBOARD_PATH.read_text(encoding="utf-8"))
    funding = json.loads(FUNDING_PATH.read_text(encoding="utf-8"))
    signals = json.loads(SIGNALS_PATH.read_text(encoding="utf-8"))
    borrowers = borrower_universe(dashboard)
    statuses: list[dict[str, Any]] = []

    sec_bdc, new_status = run_component("sec_bdc_monthly", "SEC monthly BDC data sets", "SEC filings", SEC_PAGES["bdc"], ingest_sec_bdc)
    statuses.extend(new_status)
    form_d, new_status = run_component("sec_form_d", "SEC Form D bulk data", "Entity resolution", SEC_PAGES["form_d"], ingest_form_d, borrowers)
    statuses.extend(new_status)
    etf_holdings, new_status = run_component("etf_holdings", "ETF daily holdings", "Daily loan marks", ETF_SOURCES["SRLN"]["page"], ingest_etf_holdings, borrowers)
    statuses.extend(new_status)
    gleif, new_status = run_component("gleif", "GLEIF LEI API", "Entity resolution", "https://www.gleif.org/en/lei-data/gleif-api", ingest_gleif, borrowers)
    statuses.extend(new_status)
    registries, new_status = run_component("company_registries", "Company registries", "Entity resolution", "https://developer.company-information.service.gov.uk/", ingest_company_registries, borrowers)
    statuses.extend(new_status)
    openfigi, new_status = run_component("openfigi", "OpenFIGI mapping API", "Instrument identity", "https://www.openfigi.com/api", ingest_openfigi, funding)
    statuses.extend(new_status)
    market, new_status = run_component("massive", "Massive market data API", "Market data", "https://massive.com/docs/rest/stocks/overview", ingest_massive_market)
    statuses.extend(new_status)
    macro, new_status = run_component("fred", "Federal Reserve Economic Data", "Macro drivers", "https://fred.stlouisfed.org/", ingest_fred)
    statuses.extend(new_status)
    insiders, new_status = run_component("sec_insiders", "SEC insider-transactions bulk data", "Ownership and governance", SEC_PAGES["insider"], ingest_insiders)
    statuses.extend(new_status)
    heavy_13f = args.heavy_13f or os.environ.get("BDC_ENABLE_HEAVY_13F") == "1"
    ownership, new_status = run_component("sec_13f", "SEC Form 13F bulk data", "Institutional ownership", SEC_PAGES["13f"], ingest_13f, heavy_13f)
    statuses.extend(new_status)
    legal, new_status = run_component("courtlistener", "CourtListener search API", "Legal events", "https://www.courtlistener.com/help/api/rest/", ingest_courtlistener, signals)
    statuses.extend(new_status)

    status_counts: dict[str, int] = defaultdict(int)
    for row in statuses:
        status_counts[row["status"]] += 1
    payload = {
        "meta": {
            "generated_at_utc": utc_now(),
            "pipeline_version": 1,
            "dashboard_latest_period": dashboard.get("meta", {}).get("latest_common_period"),
            "borrower_universe_count": len(borrowers),
            "status_counts": dict(sorted(status_counts.items())),
            "promotion_rule": "Only source-direct observations with an explicit as-of date may replace a dashboard market input. Entity, legal, Form D, debt-note, and name-search results remain candidates until reviewed.",
        },
        "source_status": statuses,
        "sec_bdc": sec_bdc,
        "etf_holdings": etf_holdings,
        "entity_resolution": {
            "form_d": form_d,
            "gleif": gleif,
            "registries": registries,
            "openfigi": openfigi,
        },
        "market": market,
        "macro": macro,
        "insiders": insiders,
        "institutional_ownership": ownership,
        "legal": legal,
        "limitations": [
            "A successful refresh means the provider responded and rows were parsed; it does not mean every borrower or event is covered.",
            "Daily ETF files can mix loans, cash, derivatives, and other assets. An implied mark is calculated only where the file exposes both market value and par/shares, and exact-name matches remain conservative.",
            "Form D, GLEIF, Companies House, OpenCorporates, CourtListener, and issuer-name 13F matches are candidate links, not canonical identities.",
            "SEC monthly BDC numeric and text facts are raw XBRL candidates. They do not adjust an outstanding-debt series until maturity, issuance, tender, repurchase, and redemption evidence reconciles.",
            "Credential-required feeds never report zero events when they were not queried; their state remains visibly credential_required.",
            "PACER is not queried because it can incur fees. CourtListener is the implemented free legal-alert layer and links back to reviewable dockets.",
        ],
    }
    OUTPUT_PATH.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH.relative_to(PROJECT_ROOT)}")
    print("Source status: " + ", ".join(f"{key}={value}" for key, value in sorted(status_counts.items())))
    for row in statuses:
        print(f"{row['id']}: {row['status']} ({row['records']} records) - {row['message']}")


if __name__ == "__main__":
    main()
